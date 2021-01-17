from docxtpl import DocxTemplate
import jinja2
import openpyxl
import pathlib
from pick import pick
import glob
from datetime import datetime

jinja_env = jinja2.Environment()

dir = pathlib.Path().absolute()


# Find Spreadsheet
for i in glob.glob("00.0 - DATA*"):
    data = i

# User selects AO
title = "Select the AO you wish to produce a document for:"
options = ["AO 1", "AO 2", "AO 3", "AO 4", "AO 5"]
selected_ao, index = pick(options, title)
print(selected_ao)
#print(index)
for f in glob.glob(selected_ao+"*"):
    folder = f

# User selects document
title = "Select the document you wish to produce:"
options = ["01 - Letter of Appointment BO", "01 - Letter of Appointment AO", "01 - Notices Served (ALL) - Corr (Email)", "01 - Notices Served (ALL) - Corr", "01 - Notices Served (ALL) - Prop (Email)", "01 - Notices Served (ALL) - Prop", "12.1 - Award 7th", "12.1 - Award D1", "14.1 - Schedule of Condition"]
selected_doc, index = pick(options, title)
print(selected_doc)
#print(index)

#data = '00.0 - DATA Pemberton.Steven_Caroline - 54 The Chase.xlsx'
wb = openpyxl.load_workbook(filename=data, data_only=True)

# Print(data)
ws = wb[selected_ao]

# Document variables
ao_letter_names = ws['A2'].value
ao_notice_names = ws['B2'].value
ao_fhlh = ws['C2'].value
ao_I_We = ws['D2'].value
ao_my_our = ws['E2'].value
ao_his_her = ws['F2'].value
ao_he_she = ws['G2'].value
ao_do_does = ws['H2'].value
ao_have_has = ws['I2'].value
ao_him_her = ws['J2'].value
ao_property_add_horz = ws['K2'].value
ao_correspond_add_horz = ws['L2'].value
ao_property_add_vert = ws['M2'].value
ao_correspond_add_vert = ws['N2'].value


notice_date = ws['O2'].value
#notice_date = datetime(notice_date)



bo_letter_names = ws['P2'].value
bo_I_We = ws['Q2'].value
bo_my_our = ws['R2'].value
bo_his_her = ws['S2'].value
bo_he_she = ws['T2'].value
bo_do_does = ws['U2'].value
bo_have_has = ws['V2'].value
bo_i_we = ws['W2'].value
bo_me_us = ws['X2'].value
bo_is_are = ws['Y2'].value
bo_notice_names = ws['Z2'].value
bo_correspond_add_horz = ws['AA2'].value
bo_property_add_horz = ws['AB2'].value
bo_fhlh = ws['AC2'].value
bo_correspond_add_vert = ws['AD2'].value
ao_plural = ws['AE2'].value
bo_plural = ws['AF2'].value
bo_him_her = ws['AG2'].value
ao_me_us = ws['AH2'].value
ao_myself_ourselves = ws['AI2'].value
ao_i_we = ws['AK2'].value
ao_i_amamnot_we_arearenot = ws['AL2'].value
ao_am_amnot_are_arenot = ws['AM2'].value
ao_am_are = ws['AN2'].value
s_15_12_text = ws['AO2'].value
s_15_12 = ws['AP2'].value
s1_pw_pfw = ws['AQ2'].value
s1_detail_1 = ws['AR2'].value
s1_detail_2 = ws['AS2'].value
s2_sections = ws['AU2'].value
s2_detail_1 = ws['AV2'].value
s2_detail_2 = ws['AW2'].value
s2_detail_3 = ws['AX2'].value
s2_detail_4 = ws['AY2'].value
s2_detail_5 = ws['AZ2'].value
s2_detail_6 = ws['BA2'].value
s2_detail_7 = ws['BB2'].value
s2_detail_8 = ws['BC2'].value
s2_detail_9 = ws['BD2'].value
s6_underpin_is_isnot = ws['BF2'].value
s6_sections = ws['BG2'].value
s6_detail_1 = ws['BH2'].value
s6_detail_2 = ws['BI2'].value
bo_neighbour_plural = ws['BJ2'].value
bo_apostrophe = ws['BK2'].value
ao_apostrophe = ws['BL2'].value
bo_owners_plural = ws['BN2'].value
ao_owners_plural = ws['BO2'].value
ao_is_are = ws['BP2'].value
bo_choose_plural = ws['BQ2'].value
bo_exercise_plural = ws['BR2'].value
bo_require_plural = ws['BS2'].value
ten_day_letter_subject = ws['BT2'].value
ao_choose_plural = ws['BV2'].value
ao_exercise_plural = ws['BW2'].value
ao_require_plural = ws['BX2'].value
date_appoint_surveyor_by = ws['BY2'].value
todays_date = ws['BZ2'].value
ao_i_we2 = ws['CA2'].value
worksheet_url = ws['CB2'].value
worksheet_name = ws['CC2'].value
ao_surveyor = ws['CD2'].value
aos_add_horz = ws['CE2'].value
aos_add_vert = ws['CF2'].value
AOS_gender = ws['CG2'].value
third_surveyor = ws['CH2'].value
ts_add_horz = ws['CI2'].value
ts_add_vert = ws['CJ2'].value
ts_gender = ws['CK2'].value
cc1 = ws['CL2'].value
cc2 = ws['CM2'].value
cc3 = ws['CN2'].value
cc4 = ws['CO2'].value
cc5 = ws['CP2'].value
cc6 = ws['CQ2'].value
cc7 = ws['CR2'].value
cc8 = ws['CS2'].value
cc9 = ws['CT2'].value
cc10 = ws['CU2'].value
cc11 = ws['CV2'].value
cc12 = ws['CW2'].value
cc13 = ws['CX2'].value
cc14 = ws['CY2'].value
cc15 = ws['CZ2'].value
cc16 = ws['DA2'].value
bo_surveyor = ws['DB2'].value
bos_add_horz = ws['DC2'].value
bos_add_vert = ws['DD2'].value
bos_gender = ws['DE2'].value
bos_firstname = ws['DF2'].value
aos_firstname = ws['DG2'].value
ts_firstname = ws['DH2'].value
bo_dear_mrmrs = ws['DI2'].value
ao_dear_mrmrs = ws['DJ2'].value
aos_email = ws['DK2'].value
aos_tel = ws['DL2'].value
bo_name_in_letter_body = ws['DM2'].value
architect_name = ws['DO2'].value
engineer_name = ws['DP2'].value
arch_plans_ex_horz = ws['DQ2'].value
arch_plans_dem_horz = ws['DR2'].value
arch_plans_pr_horz = ws['DS2'].value
arch_plans_ex_vert = ws['DT2'].value
arch_plans_dem_vert = ws['DU2'].value
arch_plans_pr_vert = ws['DV2'].value
eng_plans_horz = ws['DW2'].value
eng_plans_vert = ws['DX2'].value

# DocxTpl variables
context = {
    'ao_letter_names': ws['A2'].value,
    'ao_notice_names': ws['B2'].value,
    'ao_fhlh': ws['C2'].value,
    'ao_I_We': ws['D2'].value,
    'ao_my_our': ws['E2'].value,
    'ao_his_her': ws['F2'].value,
    'ao_he_she': ws['G2'].value,
    'ao_do_does': ws['H2'].value,
    'ao_have_has': ws['I2'].value,
    'ao_him_her': ws['J2'].value,
    'ao_property_add_horz': ws['K2'].value,
    'ao_correspond_add_horz': ws['L2'].value,
    'ao_property_add_vert': ws['M2'].value,
    'ao_correspond_add_vert': ws['N2'].value,
    'notice_date': notice_date.strftime("%-d %B %Y"),
    'bo_letter_names': ws['P2'].value,
    'bo_I_We': ws['Q2'].value,
    'bo_my_our': ws['R2'].value,
    'bo_his_her': ws['S2'].value,
    'bo_he_she': ws['T2'].value,
    'bo_do_does': ws['U2'].value,
    'bo_have_has': ws['V2'].value,
    'bo_i_we': ws['W2'].value,
    'bo_me_us': ws['X2'].value,
    'bo_is_are': ws['Y2'].value,
    'bo_notice_names': ws['Z2'].value,
    'bo_correspond_add_horz': ws['AA2'].value,
    'bo_property_add_horz': ws['AB2'].value,
    'bo_fhlh': ws['AC2'].value,
    'bo_correspond_add_vert': ws['AD2'].value,
    'ao_plural': ws['AE2'].value,
    'bo_plural': ws['AF2'].value,
    'bo_him_her': ws['AG2'].value,
    'ao_me_us': ws['AH2'].value,
    'ao_myself_ourselves': ws['AI2'].value,
    'ao_i_we': ws['AK2'].value,
    'ao_i_amamnot_we_arearenot': ws['AL2'].value,
    'ao_am_amnot_are_arenot': ws['AM2'].value,
    'ao_am_are': ws['AN2'].value,
    's_15_12_text': ws['AO2'].value,
    's_15_12': ws['AP2'].value,
    's1_pw_pfw': ws['AQ2'].value,
    's1_detail_1': ws['AR2'].value,
    's1_detail_2': ws['AS2'].value,
    's2_sections': ws['AU2'].value,
    's2_detail_1': ws['AV2'].value,
    's2_detail_2': ws['AW2'].value,
    's2_detail_3': ws['AX2'].value,
    's2_detail_4': ws['AY2'].value,
    's2_detail_5': ws['AZ2'].value,
    's2_detail_6': ws['BA2'].value,
    's2_detail_7': ws['BB2'].value,
    's2_detail_8': ws['BC2'].value,
    's2_detail_9': ws['BD2'].value,
    's2_detail_10': ws['BE2'].value,
    's2_detail_11': ws['BF2'].value,
    's2_detail_12': ws['BG2'].value,
    's2_detail_13': ws['BH2'].value,
    's2_detail_14': ws['BI2'].value,
    's2_detail_15': ws['BJ2'].value,
    's6_underpin_is_isnot': ws['BL2'].value,
    's6_sections': ws['BM2'].value,
    's6_detail_1': ws['BN2'].value,
    's6_detail_2': ws['BO2'].value,
    'award_sections': ws['BQ2'].value,
    'bo_neighbour_plural': ws['BS2'].value,
    'bo_apostrophe': ws['BT2'].value,
    'ao_apostrophe': ws['BU2'].value,
    'bo_owners_plural': ws['BW2'].value,
    'ao_owners_plural': ws['BX2'].value,
    'ao_is_are': ws['BY2'].value,
    'bo_choose_plural': ws['BZ2'].value,
    'bo_exercise_plural': ws['CA2'].value,
    'bo_require_plural': ws['CB2'].value,
    'ten_day_letter_subject': ws['CC2'].value,
    'ao_choose_plural': ws['CE2'].value,
    'ao_exercise_plural': ws['CF2'].value,
    'ao_require_plural': ws['CG2'].value,
    'date_appoint_surveyor_by': ws['CH2'].value,
    'todays_date': ws['CI2'].value,
    'ao_i_we2': ws['CJ2'].value,
    'worksheet_url': ws['CK2'].value,
    'worksheet_name': ws['CL2'].value,
    'ao_surveyor': ws['CM2'].value,
    'aos_add_horz': ws['CN2'].value,
    'aos_add_vert': ws['CO2'].value,
    'aos_gender': ws['CP2'].value,
    'third_surveyor': ws['CQ2'].value,
    'ts_add_horz': ws['CR2'].value,
    'ts_add_vert': ws['CS2'].value,
    'ts_gender': ws['CT2'].value,
    'cc1': ws['CU2'].value,
    'cc2': ws['CV2'].value,
    'cc3': ws['CW2'].value,
    'cc4': ws['CX2'].value,
    'cc5': ws['CY2'].value,
    'cc6': ws['CZ2'].value,
    'cc7': ws['DA2'].value,
    'cc8': ws['DB2'].value,
    'cc9': ws['DC2'].value,
    'cc10': ws['DD2'].value,
    'cc11': ws['DE2'].value,
    'cc12': ws['DF2'].value,
    'cc13': ws['DG2'].value,
    'cc14': ws['DH2'].value,
    'cc15': ws['DI2'].value,
    'cc16': ws['DJ2'].value,
    'cc17': ws['DK2'].value,
    'cc18': ws['DL2'].value,
    'cc19': ws['DM2'].value,
    'cc20': ws['DN2'].value,
    'bo_surveyor': ws['DO2'].value,
    'bos_add_horz': ws['DP2'].value,
    'bos_add_vert': ws['DQ2'].value,
    'bos_gender': ws['DR2'].value,
    'bos_firstname': ws['DS2'].value,
    'aos_firstname': ws['DT2'].value,
    'ts_firstname': ws['DU2'].value,
    'bo_dear_mrmrs': ws['DV2'].value,
    'ao_dear_mrmrs': ws['DW2'].value,
    'aos_email': ws['DX2'].value,
    'aos_tel': ws['DY2'].value,
    'bo_name_in_letter_body': ws['DZ2'].value,
    'architect_name': ws['EB2'].value,
    'engineer_name': ws['EC2'].value,
    'arch_plans_ex_horz': ws['ED2'].value,
    'arch_plans_dem_horz': ws['EE2'].value,
    'arch_plans_pr_horz': ws['EF2'].value,
    'arch_plans_ex_vert': ws['EG2'].value,
    'arch_plans_dem_vert': ws['EH2'].value,
    'arch_plans_pr_vert': ws['EI2'].value,
    'eng_plans_horz': ws['EJ2'].value,
    'eng_plans_vert': ws['EK2'].value,
    'notice_1_req': ws['EN2'].value,
    'notice_2_req': ws['EP2'].value,
    'notice_6_req': ws['ER2'].value,
}

'''print(bo_property_add_horz)
print(bo_correspond_add_horz)
print(bo_correspond_add_vert)
print(ao_property_add_horz)
print(ao_property_add_vert)
print(ao_correspond_add_horz)
print(ao_correspond_add_vert)'''
print(notice_date.strftime("%d %B %Y"))
#print(type(ao_letter_names))

tpl = DocxTemplate('templates/'+selected_doc+'.docx')
tpl.render(context, jinja_env)
tpl.save(folder+'/'+selected_doc+'.docx')


#unoconv -f ('output/'+selected_doc+'.docx', 'output/'+selected_doc+'.pdf')
#unoconv -f pdf '01 - Letter of Appointment BO.docx'
#doc2pdf custom_LoA.docx
